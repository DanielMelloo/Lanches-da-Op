import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# Mocks
class MockUser:
    name = "Daniel"

class MockSector:
    name = "CAFOR"

class MockItem:
    def __init__(self, name):
        self.name = name

class MockOrderItem:
    def __init__(self, item_name, quantity, subtotal):
        self.item = MockItem(item_name)
        self.quantity = quantity
        self.subtotal = subtotal

class MockOrder:
    def __init__(self, items):
        self.user = MockUser()
        self.sector = MockSector()
        self.tax_fixed = 0.0
        self.total_general = 101.30
        self.order_items = items

def verify_logic():
    print("Initializing Workbook...")
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Pedidos"
    
    # Simulate the logic exactly as implemented in routes_admin.py
    current_row = 2
    
    # Test Case: 1 Order with 2 Items
    order = MockOrder([
        MockOrderItem("item de teste", 1, 100.00),
        MockOrderItem("item de teste", 1, 100.00)
    ])
    
    orders = [order]
    
    for order in orders:
        order_items = order.order_items
        num_items = len(order_items)
        
        start_row = current_row
        
        for idx, order_item in enumerate(order_items):
            # Solution 1 Logic
            if idx == 0:
                ws1["A" + str(current_row)] = order.user.name
                ws1["B" + str(current_row)] = order.sector.name if order.sector else "-"
                ws1["F" + str(current_row)] = order.tax_fixed
                ws1["G" + str(current_row)] = order.total_general
            
            ws1.cell(row=current_row, column=3).value = order_item.item.name
            ws1.cell(row=current_row, column=4).value = order_item.quantity
            ws1.cell(row=current_row, column=5).value = order_item.subtotal
            
            # Format currency
            ws1.cell(row=current_row, column=5).number_format = 'R$ #,##0.00'
            
            current_row += 1
        
        # Solution 2 Logic
        if num_items > 1:
            end_row = current_row - 1
            for col_idx in [1, 2, 6, 7]:
                col_letter = get_column_letter(col_idx)
                merge_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
                ws1.merge_cells(merge_range)
                
                cell = ws1.cell(row=start_row, column=col_idx)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                if col_idx in (6, 7):
                    cell.number_format = 'R$ #,##0.00'
                    
    # Verification
    print("Verifying Merge Ranges...")
    ranges = [str(r) for r in ws1.merged_cells.ranges]
    expected_ranges = ['A2:A3', 'B2:B3', 'F2:F3', 'G2:G3']
    
    missing = [r for r in expected_ranges if r not in ranges]
    if missing:
        print(f"FAILED: Missing merge ranges: {missing}")
        print(f"Found: {ranges}")
    else:
        print("SUCCESS: All expected merge ranges found.")
        
    print("Verifying Content...")
    # Top-left of merge should have value
    if ws1["A2"].value != "Daniel":
        print(f"FAILED: A2 value wrong. Got {ws1['A2'].value}")
    else:
        print("SUCCESS: A2 has correct value.")

    # Accessing merged inner cell should be None or irrelevant, but ensuring we didn't write to it
    # Although openpyxl might create it on access, the value should be None effectively for the merge?
    # Actually, we deliberately DID NOT write to A3 in the loop (idx=0 check).
    
    print("Verifying Formatting...")
    if ws1["G2"].number_format != 'R$ #,##0.00':
        print(f"FAILED: G2 format wrong. Got {ws1['G2'].number_format}")
    else:
        print("SUCCESS: Post-merge formatting applied correctly.")

    print("Done.")

if __name__ == "__main__":
    verify_logic()
