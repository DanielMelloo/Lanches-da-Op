from flask import Blueprint, render_template, redirect, url_for, request, flash, session, current_app
from flask_login import login_required, current_user
from models import Order, Item, Subsite, Status, User, Store, Sector, OrderItem, get_sp_time
from database import db
import os, pytz
from datetime import datetime
import urllib.parse
from werkzeug.utils import secure_filename

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')

@admin_bp.before_request
@login_required
def require_admin():
    if current_user.role not in ['admin', 'admin_master']:
        flash('Acesso negado.', 'error')
        return redirect(url_for('index'))

@admin_bp.route('/dashboard')
def dashboard():
    subsite_id = current_user.subsite_id
    
    if current_user.role == 'admin_master':
        subsite_id = session.get('master_subsite_id')
        if not subsite_id:
             return redirect(url_for('master.dashboard'))
    
    if not subsite_id:
        flash('Admin sem subsite vinculado.', 'error')
        return redirect(url_for('index'))

    subsite = Subsite.query.get_or_404(subsite_id)
    recent_orders = Order.query.filter_by(subsite_id=subsite_id).order_by(Order.created_at.desc()).limit(10).all()
    
    return render_template('admin_dashboard.html', subsite=subsite, recent_orders=recent_orders)

@admin_bp.route('/orders')
def orders():
    subsite_id = current_user.subsite_id
    
    if current_user.role == 'admin_master':
        subsite_id = session.get('master_subsite_id')
        if not subsite_id: 
            return redirect(url_for('master.dashboard'))

    if not subsite_id: 
        return redirect(url_for('index'))
    
    status_filter = request.args.get('status_id')
    
    query = Order.query.filter_by(subsite_id=subsite_id)
    if status_filter:
        query = query.filter_by(status_id=status_filter)
        
    all_orders = query.order_by(Order.created_at.desc()).all()
    statuses = Status.query.all()
    
    return render_template('admin_orders.html', orders=all_orders, statuses=statuses, current_status=status_filter)

@admin_bp.route('/order/<int:order_id>/update_status', methods=['POST'])
def update_status(order_id):
    order = Order.query.get_or_404(order_id)
    new_status_id = request.form.get('status_id')
    
    if new_status_id:
        order.status_id = int(new_status_id)
        db.session.commit()
        flash('Status atualizado!', 'success')
    
    return redirect(url_for('admin.orders'))

@admin_bp.route('/order/<int:order_id>/details')
def order_details(order_id):
    order = Order.query.get_or_404(order_id)
    statuses = Status.query.all()
    sectors = Sector.query.all()
    
    # Filter items by the order's subsite only
    items = Item.query.join(Store).filter(Store.subsite_id == order.subsite_id).all()
    
    return render_template('admin_order_details.html', order=order, statuses=statuses, sectors=sectors, items=items)

@admin_bp.route('/order/<int:order_id>/delete', methods=['POST'])
def delete_order(order_id):
    order = Order.query.get_or_404(order_id)
    
    # Capture state BEFORE delete
    subsite_id = order.subsite_id
    subsite = order.subsite
    tax_mode = subsite.tax_mode if subsite else None
    
    db.session.delete(order)
    db.session.commit()
    
    # Recalculate taxes if variable mode
    if tax_mode == 'variable':
         from services.tax_service import recalculate_taxes
         recalculate_taxes(subsite_id)
         
    flash('Pedido excluído.', 'success')
    return redirect(url_for('admin.orders'))

@admin_bp.route('/orders/delete-all', methods=['POST'])
def delete_all_orders():
    subsite_id = current_user.subsite_id
    if current_user.role == 'admin_master':
        subsite_id = session.get('master_subsite_id')
    
    if not subsite_id:
        return redirect(url_for('index'))
    
    # Delete all orders for this subsite
    Order.query.filter_by(subsite_id=subsite_id).delete()
    db.session.commit()
    
    # Recalculate taxes (will be 0 since no orders, but good to reset)
    from models import Subsite
    subsite = Subsite.query.get(subsite_id)
    if subsite and subsite.tax_mode == 'variable':
         from services.tax_service import recalculate_taxes
         recalculate_taxes(subsite_id)
         
    flash('Todos os pedidos foram excluídos.', 'success')
    return redirect(url_for('admin.orders'))

@admin_bp.route('/orders/export')
def export_orders():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    from flask import send_file, redirect, url_for, flash, request, session
    import io

    # ================== CONTEXTO ==================
    subsite_id = current_user.subsite_id
    if current_user.role == 'admin_master':
        subsite_id = session.get('master_subsite_id')

    if not subsite_id:
        return redirect(url_for('index'))

    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not start_date or not end_date:
        flash('Datas não fornecidas.', 'error')
        return redirect(url_for('admin.orders'))

    start = datetime.strptime(start_date, '%Y-%m-%d')
    end = datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59)

    orders = Order.query.filter(
        Order.subsite_id == subsite_id,
        Order.created_at >= start,
        Order.created_at <= end
    ).order_by(Order.id).all()

    if not orders:
        flash('Nenhum pedido encontrado neste período.', 'warning')
        return redirect(url_for('admin.orders'))

    # ================== STYLES ==================
    header_fill = PatternFill("solid", fgColor="00B4D8")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ================== WORKBOOK ==================
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos"

    headers = ["Nome", "Setor", "Item", "Quantidade", "Total Parcial", "Taxa", "Total Geral"]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border

    current_row = 2

    # ================== DATA ==================
    for order in orders:
        items = order.order_items
        if not items:
            continue

        start_row = current_row

        for item in items:
            ws.cell(row=current_row, column=3, value=item.item.name if item.item else "Item Deletado")
            ws.cell(row=current_row, column=4, value=item.quantity)
            ws.cell(row=current_row, column=5, value=item.subtotal)

            ws.cell(row=current_row, column=5).number_format = 'R$ #,##0.00'

            for col in (3, 4, 5):
                ws.cell(row=current_row, column=col).alignment = center
                ws.cell(row=current_row, column=col).border = border

            current_row += 1

        end_row = current_row - 1

        # ===== ESCREVE APENAS NA PRIMEIRA LINHA =====
        ws.cell(start_row, 1, order.user.name if order.user else "Usuário Deletado")
        ws.cell(start_row, 2, order.sector.name if order.sector else "-")
        ws.cell(start_row, 6, order.tax_fixed)
        ws.cell(start_row, 7, order.total_general)

        ws.cell(start_row, 6).number_format = 'R$ #,##0.00'
        ws.cell(start_row, 7).number_format = 'R$ #,##0.00'

        # ===== LIMPA AS OUTRAS CÉLULAS (CRÍTICO) =====
        for r in range(start_row + 1, end_row + 1):
            for col in (1, 2, 6, 7):
                ws.cell(row=r, column=col).value = None

        # ===== MERGE + STYLE =====
        for col in (1, 2, 6, 7):
            col_letter = get_column_letter(col)
            rng = f"{col_letter}{start_row}:{col_letter}{end_row}"

            if start_row != end_row:
                ws.merge_cells(rng)

            cell = ws.cell(start_row, col)
            cell.alignment = center
            cell.border = border

    # ================== COLUMN WIDTH ==================
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # ================== SECOND SHEET ==================
    ws2 = wb.create_sheet("Por Setor")

    sector_data = {}
    all_items = set()

    for order in orders:
        sector = order.sector.name if order.sector else "Sem Setor"
        sector_data.setdefault(sector, {})

        for item in order.order_items:
            name = item.item.name if item.item else "Item Deletado"
            all_items.add(name)
            sector_data[sector][name] = sector_data[sector].get(name, 0) + item.quantity

    sectors = sorted(sector_data.keys())
    items = sorted(all_items)

    ws2.append(["Item"] + sectors)

    for col in range(1, len(sectors) + 2):
        c = ws2.cell(1, col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center

    for item in items:
        row = [item] + [sector_data[s].get(item, "") for s in sectors]
        ws2.append(row)

    ws2.column_dimensions['A'].width = 30
    for i in range(2, len(sectors) + 2):
        ws2.column_dimensions[get_column_letter(i)].width = 15

    # ================== RESPONSE ==================
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"pedidos_{start_date}_a_{end_date}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@admin_bp.route('/stores', methods=['GET', 'POST'])
def stores():
    subsite_id = current_user.subsite_id
    if current_user.role == 'admin_master':
        subsite_id = session.get('master_subsite_id')
    if not subsite_id: 
        return redirect(url_for('index'))

    if request.method == 'POST':
        action = request.form.get('action')
        store_id = request.form.get('store_id')
        name = request.form.get('name')
        whatsapp_number = request.form.get('whatsapp_number')
        
        if store_id:
            store = Store.query.get(store_id)
            if store and store.subsite_id == subsite_id:
                if action == 'toggle_active':
                    store.active = not store.active
                    db.session.commit()
                    flash(f'Loja {"ativada" if store.active else "desativada"}.', 'success')
                elif action == 'delete':
                    db.session.delete(store)
                    db.session.commit()
                    flash('Loja excluída.', 'success')
                elif action == 'manual_dispatch':
                    if not store.whatsapp_number:
                        flash(f'Loja {store.name} não possui WhatsApp cadastrado.', 'warning')
                        return redirect(url_for('admin.stores'))

                    # Signal the background dispatcher
                    store.pending_manual_dispatch = True
                    db.session.commit()
                    
                    flash(f'Disparo manual para {store.name} solicitado ao sistema.', 'success')
                    return redirect(url_for('admin.stores'))
                else:
                    # General update (name and/or whatsapp)
                    store.name = name
                    store.whatsapp_number = whatsapp_number
                    
                    # Update template if provided
                    template = request.form.get('whatsapp_template')
                    if template is not None:
                        store.whatsapp_template = template
                        
                    db.session.commit()
                    flash('Dados da loja atualizados.', 'success')
        else:
            # Create new store
            new_store = Store(
                name=name, 
                whatsapp_number=whatsapp_number,
                subsite_id=subsite_id, 
                active=True
            )
            db.session.add(new_store)
            db.session.commit()
            flash('Loja criada.', 'success')
            
        return redirect(url_for('admin.stores'))

    stores = Store.query.filter_by(subsite_id=subsite_id).all()
    return render_template('admin_stores.html', stores=stores)

    flash('Loja excluída.', 'success')
    return redirect(url_for('admin.stores'))

@admin_bp.route('/sectors', methods=['GET', 'POST'])
def sectors():
    subsite_id = current_user.subsite_id
    if current_user.role == 'admin_master':
        subsite_id = session.get('master_subsite_id')
    if not subsite_id: 
        return redirect(url_for('index'))

    if request.method == 'POST':
        action = request.form.get('action')
        sector_id = request.form.get('sector_id')
        name = request.form.get('name')
        
        if sector_id:
            sector = Sector.query.get(sector_id)
            if sector and sector.subsite_id == subsite_id:
                if action == 'toggle_active':
                    sector.active = not sector.active
                    db.session.commit()
                    flash(f'Local {"ativado" if sector.active else "desativado"}.', 'success')
                elif action == 'delete':
                    try:
                        # Try Hard Delete first (for clear DB)
                        db.session.delete(sector)
                        db.session.commit()
                        flash('Local excluído permanentemente.', 'success')
                    except Exception as e:
                        db.session.rollback()
                        # Fallback to Soft Delete if FK violation (has orders)
                        sector.active = False
                        db.session.commit()
                        flash('Local em uso foi arquivado (Soft Delete) para manter histórico.', 'warning')
                elif action == 'update_name':
                    sector.name = name
                    db.session.commit()
                    flash('Local atualizado.', 'success')
        else:
            # Create new sector
            if name:
                new_sector = Sector(name=name, subsite_id=subsite_id, active=True, type='location')
                db.session.add(new_sector)
                db.session.commit()
                flash('Local criado.', 'success')
            
        return redirect(url_for('admin.sectors'))

    # Only show locations (tables), hide menu categories created by scraper
    sectors = Sector.query.filter(
        (Sector.subsite_id == subsite_id) & 
        ((Sector.type == 'location') | (Sector.type == None))
    ).all()
    return render_template('admin_sectors.html', sectors=sectors)

@admin_bp.route('/items', methods=['GET', 'POST'])
def items():
    subsite_id = current_user.subsite_id
    if current_user.role == 'admin_master':
        subsite_id = session.get('master_subsite_id')
    if not subsite_id:
        return redirect(url_for('index'))

    if request.method == 'POST':
        item_id = request.form.get('item_id')
        name = request.form.get('name')
        price = request.form.get('price')
        store_id = request.form.get('store_id')
        image_url = request.form.get('image_url', '').strip()
        
        # Handle sub-items JSON
        import json
        subitems_json = None
        subitems_raw = request.form.get('subitems_json', '').strip()
        if subitems_raw:
            try:
                subitems_json = json.loads(subitems_raw)
            except:
                subitems_json = None
        
        # Handle file upload
        if 'image_file' in request.files:
            file = request.files['image_file']
            if file and file.filename:
                filename = secure_filename(file.filename)
                upload_dir = os.path.join(current_app.root_path, 'static', 'uploads', 'items')
                os.makedirs(upload_dir, exist_ok=True)
                filepath = os.path.join(upload_dir, filename)
                file.save(filepath)
                image_url = f'/static/uploads/items/{filename}'
        
        if item_id:
            item = Item.query.get(item_id)
            if item:
                item.name = name
                item.price = float(price)
                item.store_id = int(store_id)
                if image_url:
                    item.image_url = image_url
                if subitems_json is not None:
                    item.subitems_json = subitems_json
                db.session.commit()
                flash('Item atualizado.', 'success')
        else:
            new_item = Item(
                name=name,
                price=float(price),
                store_id=int(store_id),
                image_url=image_url if image_url else '/static/placeholders/default.png',
                subitems_json=subitems_json,
                active=True
            )
            db.session.add(new_item)
            db.session.commit()
            flash('Item criado.', 'success')
        return redirect(url_for('admin.items'))

    stores_list = Store.query.filter_by(subsite_id=subsite_id).all()
    items_list = Item.query.join(Store).filter(Store.subsite_id == subsite_id).all()
    return render_template('admin_items.html', items=items_list, stores=stores_list)

@admin_bp.route('/item/<int:item_id>/edit', methods=['GET', 'POST'])
def edit_item(item_id):
    subsite_id = current_user.subsite_id
    if current_user.role == 'admin_master':
        subsite_id = session.get('master_subsite_id')
    
    item = Item.query.get_or_404(item_id)
    
    # Check permission
    item_store = Store.query.get(item.store_id)
    if item_store and item_store.subsite_id != subsite_id:
        return "Unauthorized", 403

    if request.method == 'POST':
        item.name = request.form.get('name')
        item.price = float(request.form.get('price'))
        item.store_id = int(request.form.get('store_id'))
        
        image_url = request.form.get('image_url', '').strip()
        if request.files.get('image_file'):
            file = request.files['image_file']
            if file and file.filename:
                # Reuse filename generation logic or simplify
                from werkzeug.utils import secure_filename
                filename = secure_filename(file.filename)
                upload_dir = os.path.join(current_app.root_path, 'static', 'uploads', 'items')
                os.makedirs(upload_dir, exist_ok=True)
                file.save(os.path.join(upload_dir, filename))
                image_url = f'/static/uploads/items/{filename}'
        
        if image_url:
            item.image_url = image_url

        # Handle JSON
        import json
        subitems_raw = request.form.get('subitems_json', '').strip()
        if subitems_raw:
            try:
                item.subitems_json = json.loads(subitems_raw)
            except:
                flash('JSON inválido.', 'error')
                # Don't save if JSON is invalid? Or save partial? Let's just warn.
        else:
            item.subitems_json = None
            
        db.session.commit()
        flash('Item atualizado com sucesso.', 'success')
        return redirect(url_for('admin.items'))

    stores = Store.query.filter_by(subsite_id=subsite_id).all()
    return render_template('admin_item_edit.html', item=item, stores=stores)

@admin_bp.route('/item/<int:item_id>/delete', methods=['POST'])
def delete_item(item_id):
    item = Item.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    flash('Item excluído com sucesso.', 'success')
    return redirect(url_for('admin.items'))

@admin_bp.route('/item/<int:item_id>/toggle', methods=['POST'])
def toggle_item(item_id):
    item = Item.query.get_or_404(item_id)
    
    subsite_id = current_user.subsite_id
    if current_user.role == 'admin_master':
        subsite_id = session.get('master_subsite_id')

    # Check permission
    item_store = Store.query.get(item.store_id)
    if item_store and item_store.subsite_id != subsite_id:
        return "Unauthorized", 403
        
    item.active = not item.active
    db.session.commit()
    return redirect(url_for('admin.items'))

@admin_bp.route('/order/<int:order_id>/configure-item/<int:item_id>', methods=['GET'])
def configure_order_item(order_id, item_id):
    import json
    # Optional: order_item_id if editing existing
    order_item_id = request.args.get('order_item_id', type=int)
    
    item = Item.query.get_or_404(item_id)
    
    existing_qty = 1
    existing_subitems = []
    
    if order_item_id:
        order_item = OrderItem.query.get(order_item_id)
        if order_item:
            existing_qty = order_item.quantity
            existing_subitems = order_item.subitems_json or []
            
    return render_template('admin_order_item_edit.html', 
                           item=item,
                           order_id=order_id,
                           order_item_id=order_item_id,
                           existing_qty=existing_qty,
                           existing_subitems=existing_subitems)

@admin_bp.route('/order/save_item', methods=['POST'])
def save_order_item_admin():
    import json
    order_id = request.form.get('order_id', type=int)
    item_id = request.form.get('item_id', type=int)
    order_item_id = request.form.get('order_item_id') # Can be empty string
    
    qty = int(request.form.get('quantity', 1))
    subitems_choice = request.form.get('subitems_choice', '')
    
    order = Order.query.get_or_404(order_id)
    item = Item.query.get_or_404(item_id)
    
    subitems_data = None
    if subitems_choice:
        try:
            subitems_data = json.loads(subitems_choice)
        except:
            pass

    # Calculate Subtotal
    unit_price = item.price
    extras_price = 0.0
    
    if subitems_data:
        for group in subitems_data:
            if isinstance(group, dict):
                 p = group.get('price', 0)
                 if p: extras_price += float(p)
                 if 'options' in group and isinstance(group['options'], list):
                     for opt in group['options']:
                         op = opt.get('price', 0)
                         oq = opt.get('qty', 1)
                         if op: extras_price += (float(op) * int(oq))
    
    final_unit_price = unit_price + extras_price
    final_subtotal = final_unit_price * qty
    
    if order_item_id:
        # Update Existing
        oi = OrderItem.query.get(order_item_id)
        if oi:
            oi.quantity = qty
            oi.subitems_json = subitems_data
            oi.subtotal = final_subtotal
            oi.price_at_moment = item.price 
            flash('Item atualizado.', 'success')
    else:
        # Create New
        new_oi = OrderItem(
            order_id=order.id,
            item_id=item.id,
            quantity=qty,
            price_at_moment=item.price,
            subtotal=final_subtotal,
            subitems_json=subitems_data
        )
        db.session.add(new_oi)
        flash('Item adicionado.', 'success')
        
    db.session.commit()
    recalculate_order_totals(order.id)
    
    return redirect(url_for('admin.order_details', order_id=order_id))

def recalculate_order_totals(order_id):
    order = Order.query.get(order_id)
    if not order: return
    
    total = 0.0
    for oi in order.order_items:
        total += oi.subtotal
    
    order.total_items = total
    
    # Tax/Fees
    subsite = order.subsite
    tax_val = 0.0
    if subsite.tax_mode == 'variable':
        tax_val = subsite.calculated_variable_tax or 0.0
    else:
        tax_val = subsite.fixed_tax_value or 0.0
    order.tax_fixed = tax_val
    
    markup = 1.013 if subsite.pass_pix_tax else 1.0
    order.total_general = (order.total_items + order.tax_fixed) * markup
    order.service_fee = order.total_general - order.total_items
    
    db.session.commit()

@admin_bp.route('/order/<int:order_id>/item/<int:item_id>/delete', methods=['POST'])
def delete_order_item(order_id, item_id):
    order_item = OrderItem.query.get_or_404(item_id)
    if order_item.order_id != order_id:
        flash('Erro de integridade.', 'error')
        return redirect(url_for('admin.orders'))
        
    db.session.delete(order_item)
    db.session.commit()
    
    # Recalculate totals
    recalculate_order_totals(order_id)
    
    flash('Item removido.', 'success')
    return redirect(url_for('admin.order_details', order_id=order_id))

@admin_bp.route('/order/<int:order_id>/update_metadata', methods=['POST'])
def update_order_metadata(order_id):
    order = Order.query.get_or_404(order_id)
    
    order.status_id = int(request.form.get('status_id'))
    order.payment_status = request.form.get('payment_status')
    
    sector_id = request.form.get('sector_id')
    if sector_id:
        order.sector_id = int(sector_id)
        
    # Manual Total Overrides
    try:
        # Only update if present (checkbox logic handled in frontend form submission? 
        # Actually user requested optional override. We can just take the values.)
        order.total_items = float(request.form.get('total_items', 0))
        order.tax_fixed = float(request.form.get('tax_fixed', 0))
        order.service_fee = float(request.form.get('service_fee', 0))
        order.total_general = float(request.form.get('total_general', 0))
    except:
        pass
        
    db.session.commit()
    flash('Dados e Totais atualizados.', 'success')
    return redirect(url_for('admin.order_details', order_id=order_id))

# ============================================================
# EFÍ CONFIGURATION ROUTES
# ============================================================

@admin_bp.route('/efi-config', methods=['GET'])
def efi_config():
    subsite_id = current_user.subsite_id
    if current_user.role == 'admin_master':
         subsite_id = session.get('master_subsite_id')
    
    if not subsite_id:
        flash('Nenhum subsite selecionado.', 'error')
        return redirect(url_for('master.dashboard'))
    
    subsite = Subsite.query.get_or_404(subsite_id)
    return render_template('admin_efi_config.html', subsite=subsite)

@admin_bp.route('/efi-config/save', methods=['POST'])
def efi_config_save():
    subsite_id = current_user.subsite_id
    if current_user.role == 'admin_master':
         subsite_id = session.get('master_subsite_id')
    
    if not subsite_id:
        flash('Nenhum subsite selecionado.', 'error')
        return redirect(url_for('master.dashboard'))
    
    subsite = Subsite.query.get_or_404(subsite_id)
    
    subsite.efi_active = 'efi_active' in request.form
    subsite.efi_mode = request.form.get('efi_mode', 'producao')
    subsite.efi_client_id = request.form.get('efi_client_id')
    subsite.efi_client_secret = request.form.get('efi_client_secret')
    subsite.efi_pix_key = request.form.get('efi_pix_key')
    
    # Handle Cert Upload
    if 'efi_cert_file' in request.files:
        file = request.files['efi_cert_file']
        if file and file.filename:
            filename = secure_filename(file.filename)
            certs_dir = os.path.join(current_app.root_path, 'certs')
            os.makedirs(certs_dir, exist_ok=True)
            file.save(os.path.join(certs_dir, filename))
            subsite.efi_cert_name = filename
            
    db.session.commit()
    flash('Configurações EFí salvas com sucesso.', 'success')
    return redirect(url_for('admin.efi_config'))

@admin_bp.route('/efi-config/ajax', methods=['POST'])
def efi_config_ajax():
    """Updates interval / auto-check instantly via AJAX"""
    try:
        data = request.get_json()
        subsite_id = current_user.subsite_id
        if current_user.role == 'admin_master':
             subsite_id = session.get('master_subsite_id')
        
        if not subsite_id:
            return {'success': False, 'message': 'No subsite selected'}
        
        subsite = Subsite.query.get_or_404(subsite_id)
        
        interval = int(data.get('interval', 30))
        enabled = data.get('enabled', True)
        
        subsite.payment_check_interval = interval
        subsite.enable_auto_check = enabled
        db.session.commit()
        
        # Reschedule Job logic
        scheduler = current_app.extensions.get('scheduler')
        if scheduler:
            from services.tasks import check_all_pending_payments
            
            if enabled:
                scheduler.add_job(
                    id='check_payments', 
                    func=check_all_pending_payments, 
                    trigger='interval', 
                    seconds=interval, 
                    replace_existing=True
                )
            else:
                # If disabled, remove job
                try:
                    scheduler.remove_job('check_payments')
                except:
                    pass
            
        return {'success': True}
    except Exception as e:
        return {'success': False, 'message': str(e)}

# ============================================================
# SCRAPER ROUTES
# ============================================================

@admin_bp.route('/scraper', methods=['GET'])
def scraper():
    subsite_id = current_user.subsite_id
    if current_user.role == 'admin_master':
         subsite_id = session.get('master_subsite_id')
    
    if not subsite_id:
        return redirect(url_for('index'))
        
    stores = Store.query.filter_by(subsite_id=subsite_id).all()
    
    selected_store_id = request.args.get('store_id')
    store = None
    config = {}
    
    if selected_store_id:
        store = Store.query.get(selected_store_id)
        if store and store.subsite_id == subsite_id:
            config = store.scraper_config or {}
        else:
            store = None # Invalid or unauthorized
            
    return render_template('admin_scraper.html', stores=stores, store=store, config=config)

@admin_bp.route('/scraper/run', methods=['POST'])
@login_required
def scraper_run():
    store_id = request.form.get('store_id')
    store = Store.query.get_or_404(store_id)
    
    # Queue the job for local worker
    store.scraper_status = 'pending'
    db.session.commit()
    
    flash(f'Solicitação enviada para o terminal local! (Status: Pending)', 'info')
    return redirect(url_for('admin.scraper', store_id=store.id))

@admin_bp.route('/scraper/schedule', methods=['POST'])
def scraper_schedule():
    from services.scraper_manager import ScraperManager
    
    store_id = request.form.get('store_id')
    store = Store.query.get_or_404(store_id)
    
    config = store.scraper_config or {}
    
    config['url'] = request.form.get('url')
    config['schedule_type'] = request.form.get('schedule_type')  # interval or fixed
    
    # Interval fields (hours:minutes:seconds)
    config['interval_hours'] = int(request.form.get('interval_hours', 0))
    config['interval_minutes'] = int(request.form.get('interval_minutes', 0))
    config['interval_seconds'] = int(request.form.get('interval_seconds', 0))
    
    # Fixed schedule fields
    config['fixed_frequency'] = request.form.get('fixed_frequency', 'daily')
    config['fixed_time'] = request.form.get('fixed_time')
    # Multiple days handling
    if config['fixed_frequency'] == 'weekly':
        # getlist returns list of strings
        config['weekly_days'] = [int(d) for d in request.form.getlist('weekly_days')]
    if config['fixed_frequency'] == 'monthly':
        config['monthly_days'] = [int(d) for d in request.form.getlist('monthly_days')]
    
    config['active'] = 'active' in request.form
    
    store.scraper_config = config
    db.session.commit()
    
    # Register/Unregister Job
    ScraperManager.schedule_scraper(store.id)
    
    flash('Configuração de agendamento salva.', 'success')
    return redirect(url_for('admin.scraper', store_id=store.id))


@admin_bp.route('/settings/payment', methods=['POST'])
@login_required
def update_payment_settings():
    if current_user.role not in ['admin', 'admin_master']:
        return redirect(url_for('index'))
        
    subsite_id = current_user.subsite_id
    if current_user.role == 'admin_master':
        subsite_id = session.get('master_subsite_id')
        
    subsite = Subsite.query.get_or_404(subsite_id)
    
    action = request.form.get('action')
    
    if action == 'toggle_require_payment':
        subsite.require_payment = not subsite.require_payment
        db.session.commit()
        
        if subsite.require_payment:
            # Move all "Confirmed" orders to "Pending"
            confirmed_status = Status.query.filter_by(name='Pedido Confirmado').first()
            pending_status = Status.query.filter_by(name='Pagamento Pendente').first()
            
            if confirmed_status and pending_status:
                orders = Order.query.filter_by(subsite_id=subsite_id, status_id=confirmed_status.id).all()
                count = 0
                for o in orders:
                    o.status = pending_status
                    o.payment_required = True
                    count += 1
                flash(f'Pagamento Obrigatório ATIVADO. {count} pedidos movidos para Pendente.', 'success')
            else:
                flash('Pagamento Obrigatório ATIVADO.', 'success')
        else:
             flash('Pagamento Obrigatório DESATIVADO.', 'success')
             
    elif action == 'toggle_pass_tax':
        subsite.pass_pix_tax = not subsite.pass_pix_tax
        db.session.commit()
        flash(f'Repasse de Taxa Pix (1.3%) {"ATIVADO" if subsite.pass_pix_tax else "DESATIVADO"}.', 'success')

    elif action == 'toggle_efi_active':
        subsite.efi_active = not subsite.efi_active
        db.session.commit()
        flash(f'Integração PIX (API) {"ATIVADA" if subsite.efi_active else "DESATIVADA"}.', 'success')
    
    db.session.commit()
    
    if subsite.tax_mode == 'variable':
        from services.tax_service import recalculate_taxes
        recalculate_taxes(subsite_id)
        
    return redirect(url_for('admin.dashboard'))

@admin_bp.route('/settings/tax_goal', methods=['POST'])
@login_required
def update_tax_goal():
    subsite_id = current_user.subsite_id
    if current_user.role == 'admin_master':
        subsite_id = session.get('master_subsite_id')
    
    subsite = Subsite.query.get_or_404(subsite_id)
    
    try:
        daily_goal = float(request.form.get('daily_goal', 0))
        
        current_settings = subsite.variable_tax_settings or {}
        current_settings['daily_goal'] = daily_goal
        
        subsite.variable_tax_settings = current_settings
        
        # Ensure mode is variable
        subsite.tax_mode = 'variable'
        db.session.commit()
        
        from services.tax_service import recalculate_taxes
        recalculate_taxes(subsite_id)
        
        flash(f'Meta diária atualizada para R$ {daily_goal:.2f}', 'success')
    except ValueError:
        flash('Valor inválido.', 'error')
        
    return redirect(url_for('admin.dashboard'))

@admin_bp.route('/update_closing_schedule', methods=['POST'])
@login_required
def update_closing_schedule():
    if current_user.role not in ['admin', 'admin_master']:
        return redirect(url_for('index'))
        
    subsite_id = current_user.subsite_id
    if current_user.role == 'admin_master':
        subsite_id = session.get('master_subsite_id')
        
    subsite = Subsite.query.get_or_404(subsite_id)
    
    action = request.form.get('action')
    
    if action == 'update_time':
        subsite.order_opening_time = request.form.get('opening_time', '08:00')
        subsite.order_closing_time = request.form.get('closing_time', '23:59')
        flash('Horário de funcionamento atualizado.', 'success')
        
    elif action == 'toggle_active':
        subsite.closing_time_active = not subsite.closing_time_active
        flash(f'Fechamento automático {"ATIVADO" if subsite.closing_time_active else "DESATIVADO"}.', 'success')
        
    elif action == 'extend':
        minutes = int(request.form.get('minutes', 0))
        if minutes > 0:
            from datetime import timedelta
            now = get_sp_time()
            
            # If already has a timer and it's in the future (localized), add to it
            base_time = now
            if subsite.temp_open_until:
                target = subsite.temp_open_until
                if target.tzinfo is None:
                    target = pytz.timezone('America/Sao_Paulo').localize(target)
                
                if target > now:
                    base_time = target
            
            subsite.temp_open_until = base_time + timedelta(minutes=minutes)
            flash(f'Loja aberta por mais {minutes} minutos.', 'success')
            
    elif action == 'clear_timer':
        subsite.temp_open_until = None
        flash('Extensão de tempo removida.', 'success')
        
    elif action == 'manual_toggle':
        # Toggle subsite active status and DISABLE auto-closing to avoid confusion
        subsite.active = not subsite.active
        subsite.closing_time_active = False
        status_str = "ABERTA" if subsite.active else "FECHADA"
        flash(f'Loja {status_str} manualmente. Fechamento automático desativado.', 'success')
        
    db.session.commit()
    return redirect(url_for('admin.dashboard'))

@admin_bp.route('/get_store_status')
@login_required
def get_store_status():
    subsite_id = current_user.subsite_id
    if current_user.role == 'admin_master':
        subsite_id = session.get('master_subsite_id')
    
    subsite = Subsite.query.get_or_404(subsite_id)
    return {
        'is_open': subsite.is_open(),
        'opening_time': subsite.order_opening_time,
        'closing_time': subsite.order_closing_time,
        'closing_time_active': subsite.closing_time_active
    }
