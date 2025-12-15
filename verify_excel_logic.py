from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

def verify_logic():
    wb = Workbook()
    ws1 = wb.active
    
    # Simulate Data
    # 2 Orders
    # Order 1: 2 items
    # Order 2: 1 item
    
    data = [
        {'id': 1, 'items': ['Item A', 'Item B'], 'user': 'User 1'},
        {'id': 2, 'items': ['Item C'], 'user': 'User 2'}
    ]
    
    current_row = 2
    
    print(f"Start Row: {current_row}")
    
    for order in data:
        items = order['items']
        num_items = len(items)
        start_row = current_row
        
        print(f"Processing Order {order['id']}. Start: {start_row}")
        
        for idx, item_name in enumerate(items):
            # Write shared info on first row only
            if idx == 0:
                ws1.cell(row=current_row, column=1).value = order['user']
                print(f"  Write Shared 'User' to ({current_row}, 1)")
            
            # Write Item info
            ws1.cell(row=current_row, column=3).value = item_name
            print(f"  Write Item '{item_name}' to ({current_row}, 3)")
            
            current_row += 1
            
        if num_items > 1:
            end_row = current_row - 1
            print(f"  Merging Order {order['id']}: Rows {start_row} to {end_row}")
            
            # Merge Column 1 (User)
            merge_range = f"A{start_row}:A{end_row}"
            ws1.merge_cells(merge_range)
            
            # Style top-left
            cell = ws1.cell(row=start_row, column=1)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            print(f"  Single Item Order {order['id']}. Formatting Row {start_row}")
            cell = ws1.cell(row=start_row, column=1)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    print("\n--- Verification ---")
    print(f"Merged Cell Ranges: {ws1.merged_cells.ranges}")
    
    # Check values
    for row in range(2, current_row):
        val_a = ws1.cell(row=row, column=1).value
        val_c = ws1.cell(row=row, column=3).value
        print(f"Row {row}: Col A='{val_a}', Col C='{val_c}'")

    print("\nCheck Merged Cell Value:")
    # For a merged cell, only top-left has value. others should be None (effectively)
    # Range A2:A3
    print(f"Cell A2 value: {ws1['A2'].value}")
    print(f"Cell A3 value: {ws1['A3'].value}")

if __name__ == "__main__":
    verify_logic()
